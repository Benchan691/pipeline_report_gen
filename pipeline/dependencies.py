import logging
import os
import sys

try:
    from docx import Document
    from docx.oxml import OxmlElement
    from docx.oxml.ns import qn
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Document = None
    OxmlElement = None
    qn = None
    load_workbook = None
    Alignment = None
    get_column_letter = None


def setup_logging():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")


def check_dependencies():
    missing = []
    if Document is None:
        missing.append("python-docx")
    if load_workbook is None:
        missing.append("openpyxl")
    if not missing:
        return
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    bundled = "/Users/chankokpan/.cache/codex-runtimes/codex-primary-runtime/dependencies/python/bin/python3"
    venv_python = os.path.join(project_root, ".venv", "bin", "python3")
    hints = [
        f"Install deps: {sys.executable} -m pip install -r requirements.txt",
        f"Or use project venv: {venv_python} cnvd_docx.py --config config.json",
        f"Or use bundled Codex Python: {bundled} cnvd_docx.py --config config.json",
    ]
    sys.exit(
        "Missing Python packages: "
        + ", ".join(missing)
        + f". Current interpreter: {sys.executable}\n"
        + "\n".join(hints)
    )
